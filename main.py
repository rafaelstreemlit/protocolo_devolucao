import streamlit as st
import psycopg2
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv
import os

# Carregar variáveis de ambiente do arquivo .env
load_dotenv()

# Obter credenciais do banco de dados a partir das variáveis de ambiente
DB_HOST = os.getenv("DB_HOST")
DB_NAME = os.getenv("DB_NAME")
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")
DB_PORT = os.getenv("DB_PORT")

# Conectar ao banco de dados PostgreSQL
conn = psycopg2.connect(
    host=DB_HOST,
    dbname=DB_NAME,
    user=DB_USER,
    password=DB_PASSWORD,
    port=DB_PORT
)
cursor = conn.cursor()

# Criar a tabela 'protocolos' se ela não existir
cursor.execute('''
    CREATE TABLE IF NOT EXISTS protocolo (
        id SERIAL PRIMARY KEY,
        rota TEXT NOT NULL,
        motorista TEXT NOT NULL,
        transportadora TEXT NOT NULL,
        pedido TEXT,
        remessa TEXT,
        nota_fiscal TEXT,
        motivo TEXT,
        data_registro DATE
    )
''')
conn.commit()

def formCreation():
    st.write('Por Favor Preencha o Protocolo')

    with st.form(key="Registration Form", clear_on_submit=True):
        rota = st.text_input('Digite o número da Rota:')
        motorista = st.text_input('Digite o nome do Motorista:')
        transportadora = st.text_input('Digite o nome da Transportadora:')
        pedido = st.text_input('Digite o número do Pedido:')
        remessa = st.text_input('Digite o número da Remessa:')
        nota_fiscal = st.text_input('Digite o número da nota fiscal:')
        motivo = st.text_input('Digite o motivo da Devolução:')
        data_registro = st.date_input('Digite a Data:')
        submit = st.form_submit_button(label='Registro')

        if submit:
            st.success('O registro foi efetuado com sucesso')
            addInfo(rota, motorista, transportadora, pedido, remessa, nota_fiscal, motivo, data_registro)

def addInfo(rota, motorista, transportadora, pedido, remessa, nota_fiscal, motivo, data_registro):
    cursor.execute('''
        INSERT INTO protocolo (rota, motorista, transportadora, pedido, remessa, nota_fiscal, motivo, data_registro)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    ''', (rota, motorista, transportadora, pedido, remessa, nota_fiscal, motivo, data_registro))
    conn.commit()
    cursor.execute("SELECT LASTVAL()")
    last_id = cursor.fetchone()[0]
    st.success(f'Devolução cadastrada com sucesso. ID do Registro: {last_id}')

def viewInfo(id):
    cursor.execute("SELECT * FROM protocolo WHERE id=%s", (id,))
    result = cursor.fetchone()
    return result

def viewAllRecords():
    cursor.execute("SELECT * FROM protocolo")
    records = cursor.fetchall()
    return records

def create_excel(data):
    # Carregar o modelo de Excel
    wb = load_workbook('/Users/Samsung/Desktop/teste/modelo_devolucao.xlsx')
    ws = wb['protocolo']

    # Preencher as células especificadas
    for entry in data:
        rota, motorista, transportadora, pedido, remessa, nf, data_registro = entry[1], entry[2], entry[3], entry[4], entry[5], entry[6], entry[8]

        def split_info(info):
            separators = ['/','-',',','.',':',';','_']
            for separator in separators:
                if separator in info:
                    return info.split(separator)
            return [info]

        pedido_list = split_info(pedido)
        remessa_list = split_info(remessa)
        nf_list = split_info(nf)

        # Preencher a primeira parte
        ws['A12'] = rota
        for idx, value in enumerate(pedido_list, start=12):
            ws.cell(row=idx, column=2).value = value
        for idx, value in enumerate(remessa_list, start=12):
            ws.cell(row=idx, column=3).value = value
        for idx, value in enumerate(nf_list, start=12):
            ws.cell(row=idx, column=4).value = value
        ws['B25'] = motorista
        ws['C27'] = transportadora
        ws['B31'] = data_registro

        # Preencher a segunda parte
        ws['I12'] = rota
        for idx, value in enumerate(pedido_list, start=12):
            ws.cell(row=idx, column=10).value = value  # J12 é a coluna 10
        for idx, value in enumerate(remessa_list, start=12):
            ws.cell(row=idx, column=11).value = value  # K12 é a coluna 11
        for idx, value in enumerate(nf_list, start=12):
            ws.cell(row=idx, column=12).value = value  # L12 é a coluna 12
        ws['J25'] = motorista
        ws['K27'] = transportadora
        ws['I31'] = data_registro

    file_path = 'Protocolo_de_Entregas_Atualizado.xlsx'
    wb.save(file_path)
    return file_path

def export_to_excel(data):
    wb = Workbook()
    ws = wb.active
    ws.append(['ID', 'Rota', 'Motorista', 'Transportadora', 'Pedido', 'Remessa', 'Nota Fiscal', 'Motivo', 'Data'])
    
    for row in data:
        ws.append(row)

    file_path = 'Registros_Exportados.xlsx'
    wb.save(file_path)
    return file_path

def delete_all_records(password):
    correct_password = os.getenv("DELETE_PASSWORD")  # Obter senha do .env
    if password == correct_password:
        cursor.execute("DELETE FROM protocolo")
        conn.commit()
        st.success("Todos os registros foram excluídos com sucesso.")
    else:
        st.error("Senha incorreta. Não foi possível excluir os registros.")

def main():
    st.title("Protocolo de Devolução/Reentrega ")

    st.header("Registro de Protocolo")
    formCreation()

    st.header("Consultar Protocolo por ID")
    id_input = st.text_input("Digite o ID do protocolo:")

    if st.button("Consultar"):
        if id_input:
            try:
                id_value = int(id_input)
                result = viewInfo(id_value)
                if result:
                    st.write(f"ID: {result[0]}")
                    st.write(f"Rota: {result[1]}")
                    st.write(f"Motorista: {result[2]}")
                    st.write(f"Transportadora: {result[3]}")
                    st.write(f"Pedido: {result[4]}")
                    st.write(f"Remessa: {result[5]}")
                    st.write(f"Nota Fiscal: {result[6]}")
                    st.write(f"Data: {result[8]}")

                    data = [result]
                    file_path = create_excel(data)

                    with open(file_path, "rb") as file:
                        st.download_button(
                            label="Exportar registro para Excel",
                            data=file,
                            file_name="Protocolo_de_Entregas_Atualizado.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                else:
                    st.write("Nenhum registro encontrado para o ID fornecido.")
            except ValueError:
                st.write("Por favor, insira um ID válido.")
        else:
            st.write("Digite um ID para consultar.")

    st.header("Mostrar Todos os Registros ")
    if st.button("Mostrar Todos"):
        records = viewAllRecords()
        if records:
            for record in records:
                st.write(f"ID: {record[0]}, Rota: {record[1]}, Motorista: {record[2]}, Transportadora: {record[3]}, Pedido: {record[4]}, Remessa: {record[5]}, Nota Fiscal: {record[6]}, Data: {record[8]}")

            file_path = export_to_excel(records)

            with open(file_path, "rb") as file:
                st.download_button(
                    label="Exportar Todos para Excel",
                    data=file,
                    file_name="Registros_Exportados.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.write("Nenhum registro encontrado no banco de dados.")

    st.header("Excluir Todos os Registros")
    password_input = st.text_input("Digite a senha para excluir todos os registros:", type="password")
    if st.button("Excluir Todos os Registros"):
        delete_all_records(password_input)

if __name__ == "__main__":
    main()
